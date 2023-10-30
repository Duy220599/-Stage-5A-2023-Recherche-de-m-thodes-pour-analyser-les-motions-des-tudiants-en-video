# -*- coding: utf-8 -*-
"""
Created on Mon Apr 17 15:27:47 2023

@author: thanh
"""

import pandas as pd
import numpy as np
import os
import datetime
from scipy.stats import pearsonr
import xlsxwriter
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import matplotlib.pyplot as plt
import random
"Test parent path"
test_dir = r'E:\FileHistory'
names_list = []
figures_euclidean = {}
figures_pearsonr = {}
# print(os.listdir(os.path.dirname(test_dir)))
#Get name of people from outer folder of current dir
while len(names_list) <= 4:
    for filename in os.listdir(os.path.dirname(test_dir)):
            if not filename.endswith('.png'):
                names_list.append(filename)


def Save_graph(parent_dir):
    #Get the name of the person 
    folder_name = os.path.basename(parent_dir)
    
    #Get the session name
    session_name = os.path.basename(os.path.dirname(parent_dir))
    
    #Get the date of the session
    date = os.path.basename(os.path.dirname(os.path.dirname(parent_dir)))
    
    # Convert the date string to a datetime object
    date_obj = datetime.datetime.strptime(date, '%d-%m-%Y')
    
    
    # Generate the new file name and directory path
    file_name = f"{folder_name}-{date_obj.strftime('%d-%m')}"
    if session_name.lower() == 'first session':
        file_name += '-first-ss'

    if session_name.lower() == 'second session':
        file_name += '-second-ss'
   
    #Location to save the graphs
    parent_save_path = r'E:\Projet 2023\Data\Graph_Oximeter'
    #Sub directory of the parent_save_path
    sub_dirs = os.listdir(parent_save_path)
    #Check if name of people correspond to the folder 
    if folder_name in sub_dirs:
        save_dir = os.path.join(parent_save_path,folder_name)
        save_path = os.path.join(save_dir,f'{file_name}.jpg')
        print(f'Saving graph to {save_path}')
    return save_path

def Plot_graph_pearsonr(parent_dir,similarities_pearsonr,x_values,figs_dict):
    
    #Get the name of the person 
    name_of_person = os.path.basename(parent_dir)
    
    #Get the session name
    session_name = os.path.basename(os.path.dirname(parent_dir))
    
    #Get the date of the session
    date = os.path.basename(os.path.dirname(os.path.dirname(parent_dir)))
    # Convert the date string to a datetime object
    date_obj = datetime.datetime.strptime(date, '%d-%m-%Y')
    
    legend = f"{date_obj.strftime('%d-%m')}"

    if session_name.lower() == 'first session':
        legend += '-first-ss'
    if session_name.lower() == 'second session':
        legend += '-first-ss'      
    # Generate a random color tuple with three floats between 0 and 1
    color = tuple(np.random.rand(3,))
    
    # Get the corresponding figure from the dictionary
    fig = figs_dict_dict.get(name_of_person)
    # if len(similarities_pearsonr) > 0:
    plt.plot(similarities_pearsonr, label=legend, linewidth=10, color=color)
    plt.xlim(1, len(x_values))
    # plt.scatter(x_values, similarities_pearsonr, s=200, color=color)
    plt.title(f"Pearsonr Similarities of {name_of_person}", fontsize=60, fontweight='bold')
    plt.xlabel('Time (minutes)', fontsize=40, fontweight='bold')
    plt.ylabel('Similarity', fontsize=40, fontweight='bold')
    plt.legend(loc='center left', fontsize=30, bbox_to_anchor=(1, 0.5), frameon=False)
    fig.set_size_inches(100, 20)
    plt.tight_layout(pad=3)
    plt.savefig(Save_graph(parent_dir), dpi=300)
    plt.show()
        
def Plot_graph_euclidean(parent_dir,similarities_euclidean,x_values,figs_dict):

    #Get the name of the person 
    name_of_person = os.path.basename(parent_dir)
    
    #Get the session name
    session_name = os.path.basename(os.path.dirname(parent_dir))
    
    #Get the date of the session
    date = os.path.basename(os.path.dirname(os.path.dirname(parent_dir)))
    # Convert the date string to a datetime object
    date_obj = datetime.datetime.strptime(date, '%d-%m-%Y')
    
    legend = f"{date_obj.strftime('%d-%m')}"

    if session_name.lower() == 'first session':
        legend += '-first-ss'
    if session_name.lower() == 'second session':
        legend += '-first-ss'      
    # Generate a random color tuple with three floats between 0 and 1
    color = tuple(np.random.rand(3,))
    
    # Get the corresponding figure from the dictionary
    fig = figs_dict.get(name_of_person)
    
    # if len(similarities_pearsonr) > 0:
    plt.plot(similarities_euclidean, label=legend, linewidth= 4, color=color)
    plt.xlim(1, len(x_values))
    plt.xticks(fontsize=3)
    # plt.scatter(x_values, similarities_euclidean, s=100, color=color)
    plt.title(f"Euclidean Similarities of {name_of_person}", fontsize= 10, fontweight='bold')
    plt.xlabel('Time (minutes)', fontsize= 8 , fontweight='bold')
    plt.ylabel('Similarity', fontsize= 8, fontweight='bold')
    plt.legend(loc='center left', fontsize= 5, bbox_to_anchor=(1, 0.6), frameon=False)
    fig.set_size_inches(120, 20)
    plt.tight_layout(pad=3)
    plt.savefig(Save_graph(parent_dir), dpi=300)
    plt.show()
    
def Calculate_similarities_Euclidean(parent_dir):
    # Initialize an array to store similarities
    similarities_euclidean = []
    for filename in os.listdir(parent_dir):
        if not filename.startswith("~$") and filename.endswith(".xlsx") and filename.startswith('output'): #Use startswith to avoid temporary file 
            file_path = os.path.join(parent_dir,filename)
         
            #Read excel file
            df = pd.read_excel(file_path,engine='openpyxl',sheet_name = 'Sheet1')
            
            # Initialize an array to store similarities
            
            #Name of the person
            name =  os.path.basename(parent_dir)
            
            #Loop through columns
            for i in range(0,df.shape[1]-1):
                j = i + 1
                col1 = df.iloc[1:,i]
                col2 = df.iloc[1:,j]
                similarity = (sum(pow(a-b,2) for a, b in zip(col1, col2)))**0.5
                similarities_euclidean.append(similarity)
                
    return similarities_euclidean

def Calculate_similarities_Pearsonr(parent_dir):
    
    similarities_pearsonr = []
    #▲Find the excel file path
    for filename in os.listdir(parent_dir):
        if not filename.startswith("~$") and filename.endswith(".xlsx") and filename.startswith('output'): #Use startswith to avoid temporary file 
            file_path = os.path.join(parent_dir,filename)
         
            #Read excel file
            df = pd.read_excel(file_path,engine='openpyxl',sheet_name = 'Sheet1')
            
            #Name of the person
            name =  os.path.basename(parent_dir)
            
            #Loop through columns
            for i in range(0,df.shape[1]-1):
                j = i + 1
                col1 = df.iloc[1:,i]
                col2 = df.iloc[1:,j]
                if len(col1) >= 2 and len(col2) >= 2 and len(col1) == len(col2):
                    similarity, _ = pearsonr(col1, col2)
                    similarities_pearsonr.append(similarity)
                    
    return similarities_pearsonr
            
                
            # #Check for the name of the folder
            # if name in names_list:
            #         print(name)
            #         fig_pearsonr = figures_pearsonr[name]
            #         Plot_graph_euclidean(parent_dir,similarities_euclidean,x_values,figures_pearsonr)
                  
            # if name in names_list:
            #         print(name)
            #         fig_1 = grand_figures_pearsonr[name]
                
            #         Plot_graph_pearsonr(parent_dir, similarities_pearsonr, x_values, fig_1)
        
def Write_similarities_to_sheet(parent_dir,similarities_euclidean,similarities_pearsonr):
            # Write the similarity results to Excel
            # Write to output.xlsx
            workbook_path  = os.path.join(parent_dir,'output.xlsx')
            if os.path.exists(workbook_path):
                
                workbook = openpyxl.load_workbook(workbook_path)
                #Ceate a new sheet in excel 
                new_sheet = workbook.create_sheet('Similarities')
                # # Select the sheet you want to write to
                worksheet = new_sheet
                
                # # # Get the last row of the table
                last_row = 10
                
                # # # Skip one row from the last row of the table
                next_row = last_row + 2
                
                # # # Write the label 'Pearsonr Method' on the left of the next row
                label_cell_euclidean = worksheet.cell(row=next_row, column=1)
                label_cell_euclidean.value = 'Euclidean Method'
                label_cell_euclidean.font = Font(bold = True)
                
                # # # Write the values of the array on the row below the table
                for i, value in enumerate(similarities_euclidean):
                    worksheet.cell(row=next_row + 1, column=i+1).value = value
                
                next_row = next_row + 2
                
                # # # Write the label 'Pearsonr Method' on the left of the next row
                label_cell_pearson = worksheet.cell(row=next_row, column=1)
                label_cell_pearson.value = 'Pearsonr Method'
                label_cell_pearson.font = Font(bold = True)
                
                # # # Write the values of the array on the row below 
                for i, value in enumerate(similarities_pearsonr):
                    worksheet.cell(row=next_row + 1, column=i+1).value = value
                
                # # # Save the changes to the Excel file
                workbook.save(workbook_path)
                
            workbook_path_1  = os.path.join(parent_dir,'output1.xlsx')
                          
            if os.path.exists(workbook_path_1):
                print(1)
                workbook = openpyxl.load_workbook(workbook_path_1)
                
                #Ceate a new sheet in excel 
                new_sheet = workbook.create_sheet('Similarities')
                
                # # Select the sheet you want to write to
                worksheet = new_sheet
                
                # # # Get the last row of the table
                last_row = len(df)
                
                # # # Skip one row from the last row of the table
                next_row = last_row + 2
                
                # # # Write the label 'Pearsonr Method' on the left of the next row
                label_cell_euclidean = worksheet.cell(row=next_row, column=1)
                label_cell_euclidean.value = 'Euclidean Method'
                label_cell_euclidean.font = Font(bold = True)
                
                # # # Write the values of the array on the row below the table
                for i, value in enumerate(similarities_euclidean):
                    worksheet.cell(row=next_row + 1, column=i+1).value = value
                
                next_row = next_row + 2
                
                # # # Write the label 'Pearsonr Method' on the left of the next row
                label_cell_pearson = worksheet.cell(row=next_row, column=1)
                label_cell_pearson.value = 'Pearsonr Method'
                label_cell_pearson.font = Font(bold = True)
                
                # # # Write the values of the array on the row below 
                for i, value in enumerate(similarities_pearsonr):
                    worksheet.cell(row=next_row + 1, column=i+1).value = value
                
                # # # Save the changes to the Excel file
                workbook.save(workbook_path)
            

"Test"
similarities_euclidean = Calculate_similarities_Euclidean(test_dir)
similarities_pearsonr = Calculate_similarities_Pearsonr(test_dir)
# Write_similarities_to_sheet(test_dir,similarities_euclidean,similarities_pearsonr)
# Create the scatter plot
y_values = [i for i in range(1,len(similarities_pearsonr)+1)]
plt.plot(y_values,similarities_pearsonr, color='blue')
plt.xlabel('Time (minutes) ')
plt.ylabel('Distance ')
plt.title('Pearsonr similarity of Micheal')
plt.show()
# print(Save_graph(parent_dir))